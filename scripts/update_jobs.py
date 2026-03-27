"""
PA-jobb Uppsala – Daglig uppdatering  (v6)
──────────────────────────────────────────
Changes from v5:
  • 5 sources instead of 1: Platsbanken · Indeed RSS · JobTech Dev API · ledigajobb.se · vakanser.se
  • Indeed RSS: catches jobs posted directly on Indeed before reaching Platsbanken
  • JobTech Dev API: Sweden's official open job API, covers non-Platsbanken sources
  • Deduplication by title only (first 8 words) — company name ignored (unreliable across sources)
  • verify_still_open(): checks Platsbanken API on each run to catch filled/withdrawn jobs
  • Home coordinates updated: Hjulaxelvägen 128, 74151 Uppsala (59.9650°N, 17.7150°E)
  • Distance limit: 50 km from home address (Enköping, Norrtälje, Västerås excluded)
  • Female exclusion: 30+ keywords covering title + full listing text
  • Licence priority: 🚗 KRÄVS → ⭐⭐⭐ | 🚗 Merit → ⭐⭐ boost
  • Weekend detection: helg/fredag kväll/extrajobb → ✅ Helg/kväll
  • Dashboard: docs/index.html with 3 sections (New / Still open / Closed) + distance excluded
  • Email: short digest with stats tiles + dashboard link + new jobs table
"""

import re, json, smtplib, os, math, urllib.parse, xml.etree.ElementTree as ET
from datetime import date, datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Config ────────────────────────────────────────────────────────────────────
EXCEL_FILE   = Path("PA_Jobb_Uppsala_Gunsta.xlsx")
SEEN_FILE    = Path("scripts/seen_jobs.json")
DASHBOARD    = Path("docs/index.html")
TODAY        = date.today().isoformat()
SHEET_NAME   = "📋 Lediga Jobb"
PAGES_URL    = "https://boristeuks.github.io/pa-jobb-uppsala"
HOME_LAT     = 59.9650   # Hjulaxelvägen 128, 74151 Uppsala
HOME_LON     = 17.7150
MAX_KM       = 50

# ── Location coords ───────────────────────────────────────────────────────────
LOCATION_COORDS = {
    "gunsta":(59.9650,17.7150),"hjulaxelvägen":(59.9650,17.7150),"storvreta":(59.9647,17.7103),
    "vattholma":(59.9906,17.6856),"björklinge":(60.0003,17.5353),
    "älavsjö":(59.8700,17.7200),"uppsala":(59.8586,17.6389),
    "rickomberga":(59.8500,17.5800),"eriksberg":(59.8600,17.5400),
    "almunge":(59.9297,18.0736),"sävja":(59.8200,17.6800),
    "gottsunda":(59.8300,17.6000),"stenhagen":(59.8400,17.5600),
    "vänge":(59.8453,17.5094),"länna":(59.7800,17.7600),
    "knivsta":(59.7267,17.7847),"örbyhus":(60.2500,17.7100),
    "gimo":(60.1731,18.1847),"tobo":(60.3100,17.6300),
    "tierp":(60.3426,17.5140),"tärnsjö":(60.1486,16.9356),
    "heby":(59.9256,16.8792),"märsta":(59.6231,17.8567),
    "sigtuna":(59.6178,17.7228),"håbo":(59.5694,17.5314),
    "bålsta":(59.5694,17.5314),"rimbo":(59.7467,18.3753),
    "östhammar":(60.2563,18.3700),"enköping":(59.6350,17.0769),
    "norrtälje":(59.7580,18.7061),"västerås":(59.6162,16.5528),
    "sala":(59.9211,16.6036),"stockholm":(59.3293,18.0686),
}

# ── Filter keywords ───────────────────────────────────────────────────────────
FEMALE_EXCLUDE = [
    "kvinna","tjej","flicka","dam ","tös"," hon "," hon,"," hon.",
    "hon är","hon vill","hon bor","henne "," henne,","hennes ",
    "kvinnlig assistent","kvinnliga assistenter","söker kvinna",
    "söker en kvinna","du är kvinna","du som är kvinna",
    "till kvinna","åt kvinna","hos kvinna","till en kvinna",
    "åt en kvinna","hos en kvinna","till tjej","åt tjej","hos tjej",
    "till en tjej","åt en tjej","hos en tjej","till flicka",
    "till en flicka","kvinna med ms","ms-kvinna","dotter","syster",
]
MALE_INCLUDE = [
    "kille","man "," man,"," man.","man i ","man med ","man som ",
    "till man","åt man","hos man","till en man","åt en man",
    "pojke","grabben","honom ","hans ","herr ","son ","bror ",
    "årig kille","årsåldern","ung man","killar",
    "manliga sökande","manlig assistent",
]
LICENCE_REQUIRED = [
    "körkort krävs","körkort är ett krav","körkort krav",
    "b-körkort krävs","måste ha körkort","kräver körkort",
]
LICENCE_MERIT = [
    "körkort är meriterande","körkort meriterande","körkort önskas",
    "körkort är en merit","meriterande med körkort","körkort",
]
WEEKEND_STRONG = [
    "helg","helger","helgpass","lördag","söndag","lördagar","söndagar",
    "fredag kväll","fredagskväll","kväll och helg","kväll & helg",
    "kvällar och helger","extrajobb",
]
WEEKEND_SOFT = ["kväll","kvällar","kvällspass","deltid","vid sidan"]

HEADERS = {"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/122.0.0.0 Safari/537.36"}

# ── Helpers ───────────────────────────────────────────────────────────────────
def load_seen(): return json.loads(SEEN_FILE.read_text()) if SEEN_FILE.exists() else {}
def save_seen(s): SEEN_FILE.write_text(json.dumps(s, ensure_ascii=False, indent=2))
def norm(t): return " ".join(t.lower().split())

def is_female(title, raw=""):
    text = norm(title+" "+raw)
    return any(k in text for k in FEMALE_EXCLUDE)

def is_male(title, raw=""):
    if is_female(title, raw): return False
    return any(k in norm(title+" "+raw) for k in MALE_INCLUDE)

def lic_status(title, raw=""):
    t = norm(title+" "+raw)
    if any(k in t for k in LICENCE_REQUIRED): return "required"
    if any(k in t for k in LICENCE_MERIT):    return "merit"
    return "none"

def wk_status(title, raw=""):
    t = norm(title+" "+raw)
    if any(k in t for k in WEEKEND_STRONG): return "strong"
    if any(k in t for k in WEEKEND_SOFT):   return "soft"
    return "none"

def haversine(lat2, lon2):
    R=6371; dlat=math.radians(lat2-HOME_LAT); dlon=math.radians(lon2-HOME_LON)
    a=math.sin(dlat/2)**2+math.cos(math.radians(HOME_LAT))*math.cos(math.radians(lat2))*math.sin(dlon/2)**2
    return R*2*math.asin(math.sqrt(a))

def dist_km(ort):
    o = norm(ort)
    for key,(lat,lon) in LOCATION_COORDS.items():
        if key in o or o in key:
            return round(haversine(lat,lon))
    return None

def is_expired(d):
    try: return date.fromisoformat(d[:10]) < date.today()
    except: return False

def verify_still_open(job_id: str, url: str) -> bool:
    """
    Verifies a Platsbanken job is still actively listed.
    Returns False if 404 (filled/removed), True otherwise.
    Non-Platsbanken jobs rely on source disappearance instead.
    """
    if not job_id.startswith("pb_"):
        return True
    numeric_id = job_id.replace("pb_", "")
    # Try Platsbanken API first
    try:
        api = f"https://platsbanken-api.arbetsformedlingen.se/jobs/v1/job/{numeric_id}"
        r = requests.get(api, timeout=8, headers=HEADERS, allow_redirects=True)
        if r.status_code == 404:
            print(f"  [FILLED] #{numeric_id} returned 404 on Platsbanken API")
            return False
        if r.status_code == 200:
            try:
                data = r.json()
                if data.get("removed") or data.get("status") in ("REMOVED","FILLED"):
                    return False
            except Exception:
                pass
            return True
    except Exception:
        pass
    # Fallback: check the ad page directly
    try:
        r2 = requests.get(url, timeout=10, headers=HEADERS, allow_redirects=True)
        if r2.status_code == 404:
            return False
        closed_signals = [
            "annonsen är inte längre tillgänglig",
            "jobbet är tillsatt",
            "denna annons är avpublicerad",
            "inte tillgänglig",
            "sidan finns inte",
        ]
        if any(s in r2.text.lower() for s in closed_signals):
            return False
        return True
    except Exception:
        return True  # benefit of doubt if network fails

def lic_label(t,r=""): s=lic_status(t,r); return "🚗 KRÄVS" if s=="required" else ("🚗 Merit" if s=="merit" else "–")
def wk_label(t,r=""):  s=wk_status(t,r);  return "✅ Helg/kväll" if s=="strong" else ("🟡 Möjligt" if s=="soft" else "❓")

def rate(title, ort, raw=""):
    lic=lic_status(title,raw); wk=wk_status(title,raw)
    loc=norm(title+" "+ort)
    near=any(k in loc for k in ["knivsta","almunge","storvreta","gunsta","uppsala"])
    if lic=="required": return "⭐⭐⭐"
    if lic=="merit" and wk=="strong": return "⭐⭐⭐"
    if wk=="strong" and near: return "⭐⭐⭐"
    if wk in ("strong","soft") or lic=="merit" or near: return "⭐⭐"
    return "⭐"

def sem_flag(raw, deadline, title):
    t = norm(raw+" "+title)
    if "sommar" in t or "feriejobb" in t: return "🔴","Sommarvikariat."
    if "tills vidare" in t or "löpande" in t: return "🟢","Tills vidare."
    if "behovsanst" in t or "timvikari" in t: return "🟢","Behovsanst."
    if "6 mån" in t:
        if deadline and deadline<"2026-06-01": return "🟠","Sök INNAN semestern."
        return "🟡","Diskutera semester."
    return "🟡","Kontrollera."

def days_since(date_str):
    try: return (date.today()-date.fromisoformat(date_str[:10])).days
    except: return 0

def dedup_key(title):
    """
    Deduplication key based on title only (first 8 words, normalised).
    Ignores company — ledigajobb.se and vakanser.se often extract company
    differently or not at all, so title-only is more reliable.
    """
    t = re.sub(r'\s+', ' ', re.sub(r'[^a-zåäö0-9 ]', '', norm(title)))
    return ' '.join(t.split()[:8])

# ── Source 1: Assistanskoll (Platsbanken) ─────────────────────────────────────
def fetch_assistanskoll():
    url = "https://assistanskoll.se/platsannonser-i-Uppsala-lan.html"
    try:
        r = requests.get(url, headers=HEADERS, timeout=15); r.raise_for_status()
    except Exception as e:
        print(f"[WARN] Assistanskoll failed: {e}"); return []

    soup = BeautifulSoup(r.text, "html.parser")
    jobs = []
    for li in soup.select("ul li"):
        a = li.find("a")
        if not a: continue
        title = a.get_text(strip=True)
        href  = a.get("href", "")
        text  = li.get_text(" ", strip=True)
        m = re.search(r"/annonser/(\d+)", href)
        if not m: continue
        aid = m.group(1)
        dead = re.search(r"sista ansökningsdag (\d{4}-\d{2}-\d{2})", text)
        pub  = re.search(r"Inlämnad till Arbetsförmedlingen (\d{4}-\d{2}-\d{2})", text)
        ort  = re.search(r"\(([^.]+)\. Inlämnad", text)
        # Extract company name from listing text
        comp = re.search(r"vanlig anställning\s+([^\(]+)\(", text)
        company = comp.group(1).strip() if comp else "–"
        jobs.append({
            "id":       f"pb_{aid}",
            "title":    title,
            "url":      f"https://arbetsformedlingen.se/platsbanken/annonser/{aid}",
            "deadline": dead.group(1) if dead else "",
            "pub_date": pub.group(1)  if pub  else "",
            "ort":      ort.group(1).strip() if ort else "Uppsala",
            "company":  company,
            "source":   f"Platsbanken #{aid}",
            "source_icon": "🏛️",
            "raw_text": text,
        })
    print(f"[INFO] Assistanskoll: {len(jobs)} listings")
    return jobs

# ── Source 2: ledigajobb.se (direct ATS feeds incl. Särnmark, Vivida direct) ──
# ── Source 2: Indeed RSS ──────────────────────────────────────────────────────
def fetch_indeed_rss():
    from email.utils import parsedate_to_datetime
    seen, jobs = set(), []
    for q in ["personlig+assistent+kille","personlig+assistent+man+Uppsala","personlig+assistent+pojke"]:
        url = f"https://se.indeed.com/rss?q={q}&l=Uppsala&radius=50&sort=date&fromage=30"
        try:
            r = requests.get(url, headers=HEADERS, timeout=15); r.raise_for_status()
            root = ET.fromstring(r.content)
        except Exception as e:
            print(f"[WARN] Indeed RSS ({q}): {e}"); continue
        for item in root.findall(".//item"):
            g = tag = lambda n: (item.find(n).text or "").strip() if item.find(n) is not None else ""
            guid = tag("guid") or tag("link")
            if guid in seen: continue
            seen.add(guid)
            title,link,desc,pub = tag("title"),tag("link"),tag("description"),tag("pubDate")
            if not title or not link: continue
            pub_date = ""
            try: pub_date = parsedate_to_datetime(pub).date().isoformat() if pub else ""
            except: pass
            lm = re.search(r"(?:Plats|location)[^>]*>:?\s*([^<]+)", desc, re.I)
            cm = re.search(r"(?:retag|ompany)[^>]*>:?\s*([^<]+)", desc, re.I)
            ort = lm.group(1).strip() if lm else "Uppsala"
            company = cm.group(1).strip() if cm else "–"
            raw = " ".join(re.sub(r"<[^>]+>"," ",desc).split())
            slug = re.sub(r"[^a-z0-9]","",link.lower())[-24:]
            jobs.append({"id":f"in_{slug}","title":title,"url":link,"deadline":"",
                "pub_date":pub_date,"ort":ort,"company":company,
                "source":"Indeed RSS","source_icon":"🔴","raw_text":title+" "+raw})
    print(f"[INFO] Indeed RSS: {len(jobs)} listings"); return jobs


# ── Source 3: JobTech Dev API ─────────────────────────────────────────────────
def fetch_jobtech():
    seen, jobs = set(), []
    for q in ["personlig assistent kille","personlig assistent man Uppsala"]:
        p = urllib.parse.urlencode({"q":q,"municipality-concept-id":"682U_eMz_TXR","limit":100})
        try:
            r = requests.get(f"https://jobsearch.api.jobtechdev.se/search?{p}",
                headers={**HEADERS,"Accept":"application/json"}, timeout=15)
            r.raise_for_status(); data = r.json()
        except Exception as e:
            print(f"[WARN] JobTech ({q}): {e}"); continue
        for h in data.get("hits",{}).get("hits",[]):
            jid = str(h.get("id",""))
            if not jid or jid in seen or jid.isdigit(): continue
            seen.add(jid)
            title = h.get("headline","")
            if not title: continue
            employer = h.get("employer",{}).get("name","–")
            addr = h.get("workplace_address",{})
            ort = addr.get("municipality") or addr.get("city") or "Uppsala"
            url_ = h.get("webpage_url") or f"https://arbetsformedlingen.se/platsbanken/annonser/{jid}"
            raw = re.sub(r"<[^>]+>"," ",h.get("description",{}).get("text",""))[:1500]
            jobs.append({"id":f"jt_{jid}","title":title,"url":url_,
                "deadline":(h.get("application_deadline") or "")[:10],
                "pub_date":(h.get("publication_date") or "")[:10],
                "ort":ort,"company":employer,
                "source":"JobTech API","source_icon":"🟤","raw_text":title+" "+raw})
    print(f"[INFO] JobTech API: {len(jobs)} non-Platsbanken listings"); return jobs


def fetch_ledigajobb():
    """
    ledigajobb.se aggregates Platsbanken + direct company ATS feeds.
    This catches Särnmark reachmee + Vivida direct listings
    that appear here before hitting Platsbanken.
    """
    url = "https://www.ledigajobb.se/pr/personlig-assistent-jobb/uppsala"
    try:
        r = requests.get(url, headers=HEADERS, timeout=15); r.raise_for_status()
    except Exception as e:
        print(f"[WARN] ledigajobb.se failed: {e}"); return []

    soup = BeautifulSoup(r.text, "html.parser")
    jobs = []

    # ledigajobb uses article or div cards — find all job listing links
    for item in soup.select("article, .job-listing, .job-card, li.job"):
        a = item.find("a", href=True)
        if not a: continue
        title = a.get_text(strip=True)
        href  = a.get("href", "")
        if not href.startswith("http"):
            href = "https://www.ledigajobb.se" + href
        text  = item.get_text(" ", strip=True)

        # Skip if it's a Platsbanken job (already covered by Assistanskoll)
        if "arbetsformedlingen" in href or "platsbanken" in href:
            continue

        # Generate a stable ID from URL
        slug = re.sub(r'[^a-z0-9]', '', norm(href))[-20:]
        aid  = f"lj_{slug}"

        # Extract metadata
        dead = re.search(r"(\d{4}-\d{2}-\d{2})", text)
        pub  = re.search(r"(\d{4}-\d{2}-\d{2})", text)

        # Company name — often in a span or secondary element
        comp_el = item.find(class_=re.compile(r"company|employer|arbetsgivare", re.I))
        company = comp_el.get_text(strip=True) if comp_el else "–"

        # Location
        loc_el = item.find(class_=re.compile(r"location|ort|plats", re.I))
        ort = loc_el.get_text(strip=True) if loc_el else "Uppsala"

        if not title or len(title) < 5:
            continue

        jobs.append({
            "id":        aid,
            "title":     title,
            "url":       href,
            "deadline":  dead.group(1) if dead else "",
            "pub_date":  pub.group(1)  if pub  else "",
            "ort":       ort,
            "company":   company,
            "source":    "ledigajobb.se",
            "source_icon": "🔵",
            "raw_text":  text,
        })

    print(f"[INFO] ledigajobb.se: {len(jobs)} listings")
    return jobs

# ── Source 3: vakanser.se (another aggregator with direct feeds) ──────────────
def fetch_vakanser():
    """
    vakanser.se includes Särnmark reachmee and other direct ATS postings
    for Uppsala/Uppland region.
    """
    url = "https://vakanser.se/search/?term=personlig+assistent&region=Uppsala"
    try:
        r = requests.get(url, headers=HEADERS, timeout=15); r.raise_for_status()
    except Exception as e:
        print(f"[WARN] vakanser.se failed: {e}"); return []

    soup = BeautifulSoup(r.text, "html.parser")
    jobs = []

    for item in soup.select(".job, .vacancy, article, .listing-item"):
        a = item.find("a", href=True)
        if not a: continue
        title = a.get_text(strip=True)
        href  = a.get("href","")
        if not href.startswith("http"):
            href = "https://vakanser.se" + href
        text  = item.get_text(" ", strip=True)

        # Skip Platsbanken duplicates
        if "arbetsformedlingen" in href:
            continue

        slug = re.sub(r'[^a-z0-9]','',norm(href))[-20:]
        aid  = f"vk_{slug}"

        dead = re.search(r"sista.{0,20}(\d{4}-\d{2}-\d{2})", text, re.I)
        pub  = re.search(r"publice.{0,20}(\d{4}-\d{2}-\d{2})", text, re.I)

        comp_el = item.find(class_=re.compile(r"company|employer", re.I))
        company = comp_el.get_text(strip=True) if comp_el else "–"

        loc_el = item.find(class_=re.compile(r"location|city|ort", re.I))
        ort = loc_el.get_text(strip=True) if loc_el else "Uppsala"

        if not title or len(title) < 5:
            continue

        jobs.append({
            "id":        aid,
            "title":     title,
            "url":       href,
            "deadline":  dead.group(1) if dead else "",
            "pub_date":  pub.group(1)  if pub  else "",
            "ort":       ort,
            "company":   company,
            "source":    "vakanser.se",
            "source_icon": "🟣",
            "raw_text":  text,
        })

    print(f"[INFO] vakanser.se: {len(jobs)} listings")
    return jobs

# ── Combine + deduplicate ─────────────────────────────────────────────────────
def fetch_all():
    """
    Fetch from all sources, deduplicate globally by title+company key.
    Platsbanken entries have priority; duplicates within ANY source are also removed.
    When the same job title+company appears with multiple IDs (e.g. multi-vacancy posts),
    only the first occurrence is kept.
    """
    pb_jobs = fetch_assistanskoll()  # Source 1 — Platsbanken
    in_jobs = fetch_indeed_rss()     # Source 2 — Indeed RSS
    jt_jobs = fetch_jobtech()        # Source 3 — JobTech API
    lj_jobs = fetch_ledigajobb()     # Source 4 — ledigajobb.se
    vk_jobs = fetch_vakanser()       # Source 5 — vakanser.se

    seen_keys = {}
    all_jobs  = []
    dups      = 0

    # Priority: Platsbanken → Indeed → JobTech → ledigajobb → vakanser
    for j in pb_jobs + in_jobs + jt_jobs + lj_jobs + vk_jobs:
        key = dedup_key(j["title"])
        if key in seen_keys:
            dups += 1
            print(f"[DEDUP] '{j['title'][:50]}' already seen as {seen_keys[key]}")
            continue
        seen_keys[key] = j["id"]
        all_jobs.append(j)

    print(f"[INFO] Combined: {len(all_jobs)} unique jobs ({dups} duplicates removed)")
    print(f"[INFO] 5 sources: PB={len(pb_jobs)} In={len(in_jobs)} JT={len(jt_jobs)} LJ={len(lj_jobs)} VK={len(vk_jobs)}")
    return all_jobs

# ── Filter ────────────────────────────────────────────────────────────────────
def filter_jobs(jobs):
    kept=[]; dist_excl=[]; nf=ne=nm=0
    for j in jobs:
        if j["deadline"] and is_expired(j["deadline"]): ne+=1; continue
        if is_female(j["title"],j.get("raw_text","")): nf+=1; continue
        if not is_male(j["title"],j.get("raw_text","")): nm+=1; continue
        d = dist_km(j.get("ort",""))
        if d is not None and d > MAX_KM:
            dist_excl.append({**j,"distance_km":d}); continue
        j["distance_km"] = d
        kept.append(j)
    print(f"[INFO] Filter: {len(kept)} kept | female={nf} expired={ne} no-male={nm} far={len(dist_excl)}")
    return kept, dist_excl

def enrich(j, seen):
    raw = j.get("raw_text","")
    j["stars"]      = rate(j["title"], j["ort"], raw)
    j["korkort"]    = lic_label(j["title"], raw)
    j["tider"]      = wk_label(j["title"], raw)
    j["anst"]       = "Deltid"
    sf, sn          = sem_flag(raw, j["deadline"], j["title"])
    j["sem_flag"]   = sf
    j["sem_note"]   = sn
    j["avstand"]    = f"~{j['distance_km']} km" if j.get("distance_km") else "? km"
    j["first_seen"] = seen.get(j["id"],{}).get("first_seen", TODAY)
    return j

# ── Excel ─────────────────────────────────────────────────────────────────────
MCOL={"⭐⭐⭐":("1B4332","D8F3DC"),"⭐⭐":("1A3A5C","DBEAFE"),"⭐":("4A1D1D","FEE2E2")}
SBG={"🟢":"C8E6C9","🟠":"FFE0B2","🔴":"FFCDD2","🟡":"FFF9C4"}
SFG={"🟢":"1B5E20","🟠":"E65100","🔴":"B71C1C","🟡":"F57F17"}

def tb():
    s=Side(style="thin",color="CCCCCC"); return Border(left=s,right=s,top=s,bottom=s)

def wc(ws,row,col,v,bg,fg,bold=False,align="left",wrap=True,ul=False,hl=None):
    c=ws.cell(row=row,column=col,value=v)
    c.font=Font(name="Arial",size=9,bold=bold,color=fg,underline="single" if ul else None)
    c.fill=PatternFill("solid",start_color=bg)
    c.alignment=Alignment(wrap_text=wrap,vertical="top",horizontal=align)
    c.border=tb()
    if hl: c.hyperlink=hl

def safe_unmerge(ws,rn):
    for ref in [str(m) for m in ws.merged_cells.ranges if m.min_row<=rn<=m.max_row]:
        try: ws.unmerge_cells(ref)
        except: pass

def write_row(ws,rn,j,is_new=False):
    stars=j.get("stars","⭐⭐"); hbg,rbg=MCOL[stars]
    sf=j.get("sem_flag","🟡"); sn=j.get("sem_note","")
    sbg=SBG.get(sf,"FFFFFF"); sfg=SFG.get(sf,"000000")
    src_icon = j.get("source_icon","🏛️")
    pub = f"{'🆕 NYTT'+chr(10) if is_new else ''}{j.get('pub_date','?')}\n{src_icon} {j.get('source','')}"
    safe_unmerge(ws,rn)
    wc(ws,rn,1,("🆕 " if is_new else "")+stars,hbg,"FFFFFF",True,"center")
    wc(ws,rn,2,j["title"],rbg,"1A3A8C",True,"left",True,True,j["url"])
    wc(ws,rn,3,j.get("company","–"),rbg,"000000")
    wc(ws,rn,4,j.get("ort","Uppsala"),rbg,"000000")
    wc(ws,rn,5,j.get("anst","Deltid"),rbg,"000000")
    wc(ws,rn,6,j.get("tider","❓"),rbg,"000000")
    wc(ws,rn,7,j.get("korkort","–"),rbg,"000000",False,"center")
    wc(ws,rn,8,j.get("deadline","Löpande"),rbg,"000000",False,"center")
    wc(ws,rn,9,j.get("avstand","–"),rbg,"000000",False,"center")
    wc(ws,rn,10,f"{j.get('korkort','–')} · {j.get('tider','–')}",rbg,"000000")
    wc(ws,rn,11,"♂️ Man","DBEAFE","1E40AF",True,"center")
    wc(ws,rn,12,f"{sf}\n{sn}",sbg,sfg)
    wc(ws,rn,13,pub,rbg,"333333")
    ws.row_dimensions[rn].height=55

def update_excel(new_jobs, removed_ids):
    if not EXCEL_FILE.exists(): print("[ERROR] Excel not found"); return
    wb=load_workbook(str(EXCEL_FILE))
    if SHEET_NAME not in wb.sheetnames: print("[ERROR] Sheet not found"); return
    ws=wb[SHEET_NAME]
    last=5
    for row in ws.iter_rows(min_row=5,max_row=ws.max_row):
        v=row[1].value
        if v and "BORTTAGNA" in str(v): break
        if v: last=row[0].row
    footer=last+1
    for j in new_jobs:
        ws.insert_rows(footer); write_row(ws,footer,j,is_new=True); footer+=1
    safe_unmerge(ws,1); ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=13)
    c=ws.cell(row=1,column=1,
        value=f"📋 PA-jobb Uppsala · UPPDATERAD {TODAY} · 5 källor · Radie <{MAX_KM} km · {len(new_jobs)} nya · {len(removed_ids)} borttagna")
    c.font=Font(name="Arial",size=11,bold=True,color="FFFFFF")
    c.fill=PatternFill("solid",start_color="1B4332")
    c.alignment=Alignment(horizontal="left",vertical="center")
    wb.save(str(EXCEL_FILE))
    print(f"[INFO] Excel updated")

# ── Dashboard ─────────────────────────────────────────────────────────────────
def sc(s): return {"⭐⭐⭐":"#1B4332","⭐⭐":"#1A3A5C","⭐":"#991B1B"}.get(s,"#333")

def job_card(j, badge_html=""):
    d    = j.get("distance_km"); dist=f"~{d} km" if d else "?"
    days = days_since(j.get("first_seen",TODAY))
    dl   = j.get("deadline","Löpande")
    dl_style = ' style="color:#991B1B;font-weight:600"' if dl and dl<TODAY else ""
    src_icon = j.get("source_icon","🏛️")
    src_name = j.get("source","Platsbanken")
    return f"""
    <div class="card">
      <div class="card-top">
        <span class="stars" style="color:{sc(j.get('stars','⭐'))}">{j.get('stars','⭐')}</span>
        <a href="{j['url']}" target="_blank" class="job-title">{j['title']}</a>
        {badge_html}
      </div>
      <div class="card-meta">
        <span>📍 {j.get('ort','?')} &middot; {dist}</span>
        <span>🏢 {j.get('company','–')}</span>
        <span>🚗 {j.get('korkort','–')}</span>
        <span>{j.get('tider','–')}</span>
        <span{dl_style}>⏰ {dl}</span>
        <span>{j.get('sem_flag','🟡')}</span>
        <span class="src-badge">{src_icon} {src_name}</span>
        {"<span class='days-open'>öppen "+str(days)+" dagar</span>" if days>0 else ""}
      </div>
    </div>"""

def build_dashboard(new_jobs, open_jobs, closed_jobs, dist_excl, source_stats):
    DASHBOARD.parent.mkdir(exist_ok=True)
    new_html   = "".join(job_card(j,'<span class="badge badge-new">NY IDAG</span>') for j in new_jobs) or "<p class='empty'>Inga nya jobb idag.</p>"
    open_html  = "".join(job_card(j) for j in open_jobs) or "<p class='empty'>Inga öppna jobb från tidigare dagar.</p>"
    closed_html= "".join(
        f'<div class="card card-closed"><span class="job-title-closed">{j.get("title","?")}</span>'
        f'<span class="badge badge-closed">STÄNGD</span></div>'
        for j in closed_jobs) or "<p class='empty'>Inga nyligen stängda.</p>"
    dist_rows = "".join(
        f'<div class="dist-row"><a href="{j["url"]}" target="_blank">{j["title"]}</a>'
        f' <span class="dist-badge">{j.get("ort","?")} · ~{j.get("distance_km","?")} km</span>'
        f' <span class="src-badge2">{j.get("source_icon","🏛️")} {j.get("source","?")}</span></div>'
        for j in dist_excl)
    src_html = " &nbsp;|&nbsp; ".join(f"{v} från {k}" for k,v in source_stats.items())

    html = f"""<!DOCTYPE html>
<html lang="sv">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>PA-jobb Uppsala – {TODAY}</title>
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f5f5f4;color:#1c1917;line-height:1.5}}
  header{{background:#1B4332;color:white;padding:20px 24px}}
  header h1{{font-size:20px;font-weight:600}}
  header p{{opacity:.75;font-size:12px;margin-top:4px}}
  .stats{{display:flex;gap:10px;margin-top:14px;flex-wrap:wrap}}
  .stat{{background:rgba(255,255,255,.15);border-radius:8px;padding:8px 14px;font-size:12px;font-weight:500;min-width:80px;text-align:center}}
  .stat span{{font-size:22px;font-weight:700;display:block}}
  .sources{{background:rgba(255,255,255,.1);border-radius:6px;padding:6px 12px;font-size:11px;margin-top:10px;opacity:.8}}
  main{{max-width:900px;margin:24px auto;padding:0 16px 40px}}
  .section{{margin-bottom:28px}}
  .section-title{{font-size:15px;font-weight:600;margin-bottom:12px;display:flex;align-items:center;gap:8px}}
  .card{{background:white;border:1px solid #e5e7eb;border-radius:10px;padding:14px 16px;margin-bottom:8px;transition:box-shadow .15s}}
  .card:hover{{box-shadow:0 2px 8px rgba(0,0,0,.08)}}
  .card-top{{display:flex;align-items:flex-start;gap:8px;flex-wrap:wrap}}
  .stars{{font-size:12px;font-weight:700;flex-shrink:0;padding-top:2px}}
  .job-title{{color:#1A3A5C;font-weight:600;font-size:14px;text-decoration:none;flex:1}}
  .job-title:hover{{text-decoration:underline}}
  .card-meta{{display:flex;gap:8px;margin-top:8px;font-size:12px;color:#57534e;flex-wrap:wrap;align-items:center}}
  .days-open{{color:#854F0B;background:#FEF9C3;padding:1px 7px;border-radius:99px;font-size:11px}}
  .src-badge{{background:#EDE9FE;color:#4C1D95;padding:1px 7px;border-radius:99px;font-size:11px}}
  .badge{{padding:2px 8px;border-radius:99px;font-size:11px;font-weight:600;flex-shrink:0}}
  .badge-new{{background:#D8F3DC;color:#1B4332}}
  .badge-closed{{background:#FEE2E2;color:#991B1B}}
  .card-closed{{display:flex;align-items:center;gap:10px;opacity:.6;background:white;border:1px solid #e5e7eb;border-radius:10px;padding:10px 16px;margin-bottom:6px}}
  .job-title-closed{{text-decoration:line-through;color:#57534e;font-size:13px;flex:1}}
  .dist-section{{background:#FFF5F5;border:1px solid #FECACA;border-radius:10px;padding:14px 16px;margin-bottom:28px}}
  .dist-title{{font-size:13px;font-weight:600;color:#991B1B;margin-bottom:10px}}
  .dist-row{{padding:6px 0;border-bottom:1px solid #FEE2E2;font-size:12px;display:flex;gap:8px;align-items:center;flex-wrap:wrap}}
  .dist-row:last-child{{border-bottom:none}}
  .dist-row a{{color:#1A3A5C}}
  .dist-badge{{background:#FEE2E2;color:#991B1B;padding:1px 7px;border-radius:99px;font-size:11px;white-space:nowrap}}
  .src-badge2{{background:#F3F4F6;color:#374151;padding:1px 7px;border-radius:99px;font-size:11px}}
  .empty{{color:#a8a29e;font-size:13px;padding:8px 0}}
  .legend{{background:white;border:1px solid #e5e7eb;border-radius:10px;padding:12px 16px;font-size:12px;color:#57534e;margin-top:8px;line-height:2}}
  footer{{text-align:center;font-size:11px;color:#a8a29e;padding:20px}}
</style>
</head>
<body>
<header>
  <h1>📋 PA-jobb Uppsala</h1>
  <p>Uppdaterad {TODAY} &middot; Manliga kunder &middot; Max {MAX_KM} km från Hjulaxelvägen 128 &middot; Körkort &amp; helg prioriterat</p>
  <div class="stats">
    <div class="stat"><span>{len(new_jobs)}</span>Nya idag</div>
    <div class="stat"><span>{len(open_jobs)}</span>Öppna</div>
    <div class="stat"><span>{len(closed_jobs)}</span>Stängda</div>
    <div class="stat"><span>{len(dist_excl)}</span>För långt</div>
  </div>
  <div class="sources">📡 Källor: {src_html}</div>
</header>
<main>
  <div class="section">
    <div class="section-title" style="color:#1B4332">🆕 Nya jobb idag ({len(new_jobs)})</div>
    {new_html}
  </div>
  <div class="section">
    <div class="section-title" style="color:#1A3A5C">📂 Fortfarande öppna ({len(open_jobs)})</div>
    {open_html}
  </div>
  <div class="section">
    <div class="section-title" style="color:#991B1B">🗑️ Nyligen stängda ({len(closed_jobs)})</div>
    {closed_html}
  </div>
  {"" if not dist_excl else f'<div class="dist-section"><div class="dist-title">📍 Exkluderade – för långt bort (&gt;{MAX_KM} km) ({len(dist_excl)})</div>{dist_rows}</div>'}
  <div class="legend">
    <strong>Källor:</strong> 🏛️ Platsbanken &nbsp;|&nbsp; 🔴 Indeed RSS &nbsp;|&nbsp; 🟤 JobTech API &nbsp;|&nbsp; 🔵 ledigajobb.se &nbsp;|&nbsp; 🟣 vakanser.se<br>
    <strong>Rating:</strong> ⭐⭐⭐ Perfekt &nbsp;|&nbsp; ⭐⭐ Bra &nbsp;|&nbsp; ⭐ Möjlig<br>
    <strong>Körkort:</strong> 🚗 KRÄVS = obligatoriskt &nbsp;|&nbsp; 🚗 Merit = meriterande<br>
    <strong>Helg:</strong> ✅ Helg/kväll = nämns explicit &nbsp;|&nbsp; 🟡 Möjligt = kväll/deltid nämns
  </div>
</main>
<footer>PA-agent v5 · 5 källor · GitHub Actions · Boris Teuks · {TODAY}</footer>
</body></html>"""

    DASHBOARD.write_text(html, encoding="utf-8")
    print(f"[INFO] Dashboard written → {DASHBOARD}")

# ── Email ─────────────────────────────────────────────────────────────────────
def send_email(new_jobs, open_jobs, closed_jobs, dist_excl, source_stats):
    sender=os.environ.get("EMAIL_FROM"); pw=os.environ.get("EMAIL_PASSWORD"); rcv=os.environ.get("EMAIL_TO")
    if not all([sender,pw,rcv]): print("[WARN] No email credentials"); return

    rows=""
    for j in sorted(new_jobs,key=lambda x:x.get("stars","⭐"),reverse=True):
        c={"⭐⭐⭐":"#1B4332","⭐⭐":"#1A3A5C","⭐":"#991B1B"}.get(j.get("stars","⭐"),"#333")
        d=j.get("distance_km"); dl=f"~{d} km" if d else "?"
        rows+=f"""<tr>
          <td style="padding:8px 10px;border-bottom:1px solid #eee">
            <span style="color:{c};font-weight:700;font-size:11px">{j.get('stars','⭐')}</span>
            <b> <a href="{j['url']}" style="color:#1A3A5C;text-decoration:none">{j['title']}</a></b><br>
            <span style="color:#888;font-size:11px">{j.get('ort','?')} · {dl} · {j.get('source_icon','🏛️')} {j.get('source','?')}</span>
          </td>
          <td style="padding:8px;border-bottom:1px solid #eee;text-align:center;font-size:12px">{j.get('korkort','–')}</td>
          <td style="padding:8px;border-bottom:1px solid #eee;text-align:center;font-size:12px">{j.get('tider','–')}</td>
          <td style="padding:8px;border-bottom:1px solid #eee;text-align:center;font-size:12px">{j.get('deadline','Löpande')}</td>
        </tr>"""

    src_summary = " &middot; ".join(f"{v} från {k}" for k,v in source_stats.items())
    new_section = f"""
    <table style="width:100%;border-collapse:collapse;background:white;border:1px solid #e5e7eb;border-radius:6px">
      <tr style="background:#1B4332;color:white;font-size:11px">
        <th style="padding:8px 10px;text-align:left">Annons</th>
        <th style="padding:8px">Körkort</th><th style="padding:8px">Helg</th><th style="padding:8px">Deadline</th>
      </tr>{rows}</table>""" if new_jobs else "<p style='color:#888;font-size:13px'>Inga nya jobb idag.</p>"

    html=f"""<html><body style="font-family:Arial,sans-serif;max-width:720px;margin:auto">
    <div style="background:#1B4332;color:white;padding:18px 20px;border-radius:8px 8px 0 0">
      <h2 style="margin:0;font-size:18px">📋 PA-jobb Uppsala – {TODAY}</h2>
      <p style="margin:4px 0 0;opacity:.75;font-size:12px">Daglig uppdatering · 5 källor · Manliga kunder · Max {MAX_KM} km från Hjulaxelvägen 128</p>
      <p style="margin:4px 0 0;opacity:.6;font-size:11px">📡 {src_summary}</p>
    </div>
    <div style="padding:16px 20px;background:#f9f9f9;border:1px solid #e5e7eb;border-top:none;border-radius:0 0 8px 8px">
      <div style="display:flex;gap:10px;margin-bottom:16px;flex-wrap:wrap">
        <div style="background:#D8F3DC;border-radius:8px;padding:8px 14px;text-align:center;min-width:80px">
          <div style="font-size:22px;font-weight:700;color:#1B4332">{len(new_jobs)}</div>
          <div style="font-size:11px;color:#1B4332">Nya idag</div>
        </div>
        <div style="background:#DBEAFE;border-radius:8px;padding:8px 14px;text-align:center;min-width:80px">
          <div style="font-size:22px;font-weight:700;color:#1A3A5C">{len(open_jobs)}</div>
          <div style="font-size:11px;color:#1A3A5C">Öppna</div>
        </div>
        <div style="background:#FEE2E2;border-radius:8px;padding:8px 14px;text-align:center;min-width:80px">
          <div style="font-size:22px;font-weight:700;color:#991B1B">{len(closed_jobs)}</div>
          <div style="font-size:11px;color:#991B1B">Stängda</div>
        </div>
        <div style="flex:1;background:#1B4332;border-radius:8px;padding:10px 16px;display:flex;align-items:center;justify-content:center;min-width:160px">
          <a href="{PAGES_URL}" style="color:white;font-weight:600;font-size:14px;text-decoration:none">🔗 Dashboard →</a>
        </div>
      </div>
      <h3 style="color:#1B4332;margin:0 0 10px">🆕 Nya jobb idag</h3>
      {new_section}
      <p style="margin-top:14px;font-size:12px;color:#888;text-align:center">
        Alla {len(open_jobs)} öppna jobb → <a href="{PAGES_URL}" style="color:#1A3A5C">{PAGES_URL}</a>
      </p>
      <p style="margin-top:4px;font-size:11px;color:#aaa;text-align:center">Excel bifogad · PA-agent v5 · GitHub Actions</p>
    </div></body></html>"""

    subj=(f"PA-jobb Uppsala {TODAY} – {len(new_jobs)} nya, {len(open_jobs)} öppna")
    msg=MIMEMultipart("alternative"); msg["Subject"]=subj; msg["From"]=sender; msg["To"]=rcv
    msg.attach(MIMEText(html,"html"))
    if EXCEL_FILE.exists():
        with open(EXCEL_FILE,"rb") as f:
            p=MIMEBase("application","octet-stream"); p.set_payload(f.read())
        encoders.encode_base64(p); p.add_header("Content-Disposition",f"attachment; filename={EXCEL_FILE.name}")
        msg.attach(p)
    try:
        with smtplib.SMTP_SSL("smtp.gmail.com",465) as s:
            s.login(sender,pw); s.sendmail(sender,rcv,msg.as_string())
        print(f"[INFO] Email sent to {rcv}")
    except Exception as e:
        print(f"[ERROR] Email failed: {e}")

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print(f"\n{'='*55}\nPA-agent v5 · {datetime.now().strftime('%Y-%m-%d %H:%M')}\n{'='*55}\n")

    seen     = load_seen()
    all_raw  = fetch_all()
    filtered, dist_excl = filter_jobs(all_raw)

    # Source stats for email/dashboard
    source_stats = {}
    for j in filtered:
        src = j.get("source","Platsbanken")
        # Shorten to base name
        base = "Platsbanken" if "Platsbanken" in src else src.split(".")[0]
        source_stats[base] = source_stats.get(base,0) + 1

    current_ids = {j["id"]:j for j in filtered}
    new_jobs=[]; open_jobs=[]; closed_jobs=[]

    # Count previously-seen jobs that need live verification
    n_to_verify = sum(1 for j in filtered if j["id"] in seen)
    print(f"[INFO] Verifying {n_to_verify} previously-seen jobs against live Platsbanken...")

    for j in filtered:
        enrich(j, seen)
        if j["id"] not in seen:
            j["first_seen"] = TODAY
            new_jobs.append(j)
            print(f"[NEW  {j['stars']}] {j['title'][:50]} | {j['ort']} | {j['korkort']} | {j['tider']} | {j.get('source_icon','')} {j.get('source','')}")
        else:
            # Verify still accepting applications on live site
            if verify_still_open(j["id"], j["url"]):
                open_jobs.append(j)
            else:
                closed_jobs.append({
                    "id": j["id"], "title": j["title"],
                    "url": j["url"], "closed_reason": "filled/withdrawn"
                })
                print(f"[FILLED] {j['title'][:55]}")

    for aid, meta in seen.items():
        if aid not in current_ids:
            if not any(c["id"] == aid for c in closed_jobs):
                closed_jobs.append({"id":aid,"title":meta.get("title","?"),"url":"#"})

    for lst in [new_jobs, open_jobs]:
        lst.sort(key=lambda j:j.get("stars","⭐"),reverse=True)

    seen.update({j["id"]:{"title":j["title"],"first_seen":j["first_seen"]} for j in new_jobs})
    for j in open_jobs:
        if j["id"] in seen: seen[j["id"]]["title"] = j["title"]
    for j in closed_jobs: seen.pop(j["id"],None)
    save_seen(seen)

    build_dashboard(new_jobs, open_jobs, closed_jobs, dist_excl, source_stats)
    if new_jobs or closed_jobs:
        update_excel(new_jobs, [j["id"] for j in closed_jobs])
    else:
        print("[INFO] No changes – Excel unchanged.")
    send_email(new_jobs, open_jobs, closed_jobs, dist_excl, source_stats)

    print(f"\n✅ Done! New={len(new_jobs)} Open={len(open_jobs)} Closed={len(closed_jobs)} Far={len(dist_excl)}")
    print(f"   Sources: {source_stats}")

if __name__=="__main__":
    main()
